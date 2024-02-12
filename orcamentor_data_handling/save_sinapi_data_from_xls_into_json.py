import openpyxl
import json

# Load the workbook and get the active sheet
wb = openpyxl.load_workbook('caderno.xlsx')
sheet = wb.active

# Create an empty dictionary to hold the pairs
data_pairs = {}

for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=2, values_only=True):
    # Extract the code from the first column
    code_full = row[0]

    # Split the code and get the second (XXXX) and fourth elements (CCCC) using the provided pattern
    try:
        code_part1 = code_full.split('.')[1]
        code_part2 = code_full.split('.')[2].split('/')[0].split('-')[0]
    except:
        continue

    # Convert the tuple to string and use as key
    key = f"{code_part1}_{code_part2}"

    # Check if the combined code is already in the dictionary or if the row is not useful (i.e., contains 'INSUMO' or 'COMPOSICAO')
    if key not in data_pairs and row[1] not in ['INSUMO', 'COMPOSICAO']:
        data_pairs[key] = row[1]

# Save the dictionary as a JSON file
with open('output.json', 'w') as json_file:
    json.dump(data_pairs, json_file, indent=4)

print("Extraction completed!")
