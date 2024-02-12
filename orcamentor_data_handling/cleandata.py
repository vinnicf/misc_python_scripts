import openpyxl

def clean_data_in_xlsx(file_path):
    # Load the workbook and select the first worksheet
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Iterate over the rows in the worksheet
    for row in ws.iter_rows(min_row=2, max_col=5, max_row=ws.max_row):  # Assuming header is in the first row with min_row=2
        cell = row[4]  # 5th column
        if cell.value and isinstance(cell.value, str):
            # Replace "," with "." and convert to float
            try:
                cleaned_value = float(cell.value.replace(",", "."))
                cell.value = cleaned_value
            except ValueError:
                print(f"Failed to convert value {cell.value} in row {cell.row}")

    # Save the modified workbook
    wb.save("modified_" + file_path)


clean_data_in_xlsx("no-insumos.xlsx")
